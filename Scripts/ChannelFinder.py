#! /usr/local/bin/python

import openpyxl
import re
import os
import sys
import zipfile
# This is the regex we use to identify a string to be a channel name
regex = re.compile('^[HL]1:[A-Z]{3}-.*')

# This function takes in an xlsx file (spreadsheet) and generates all
# the distinct channels in the sheet and output as a txt with the same
# name. It generates error if things goes wrong (like file not found).

def process_workbook(workbook):
    # Read in the file, set the parameters to suppress the user warnings
    wb = openpyxl.load_workbook(workbook, read_only = True, data_only = True)
    sheet = wb.get_sheet_by_name('Sheet')

    print ('opened the file! Highest-row:{}'.format(sheet.get_highest_row()))
    # Use a map from channel number to channel name to save all channel
    # names met during traverse
    channel_name = {};

    # Iterate all the cells, add it to the hashmap if it meets these conditions
    count = 0
    for row in sheet.rows:
        print ('Start row {}...'.format(count))
        for cell in row:
            if (type(cell.value).__name__ == 'unicode' and \
                re.match(regex, cell.value) and \
                sheet.cell(row = cell.row, column = cell.column - 1).value != None and \
                not (sheet.cell(row = cell.row, column = cell.column - 1).value in channel_name)):
                channel_name[sheet.cell(row = cell.row, column = cell.column - 1).value] = \
                    sheet.cell(row = cell.row, column = cell.column).value
        count = count + 1
    print ('Finished the reading of {}'.format(workbook))
    # Print out everything in the map to the a txt file with the same name
    # under the ./output folder in the spreadsheets folder
    # Note: create file / folder when not found, overwrite when exist
    output_path = workbook.rsplit('/', 1)[0] + \
        '/output/' + workbook.rsplit('/', 1)[1].rsplit('.', 1)[0] + \
        '.txt'
    output_txt = open(output_path, "wb")
    for num in channel_name:
        output_txt.write("{}\n".format(channel_name[num]))
    output_txt.close()
    print ('Finished the writing')

def zipdir(dirPath=None, zipFilePath=None, includeDirInZip=True):
    # I get this function from:
    # http://peterlyons.com/problog/2009/04/zip-dir-python
    """Create a zip archive from a directory.
    
    Keyword arguments:
    
    dirPath -- string path to the directory to archive. This is the only
    required argument. It can be absolute or relative, but only one or zero
    leading directories will be included in the zip archive.

    zipFilePath -- string path to the output zip file. This can be an absolute
    or relative path. If the zip file already exists, it will be updated. If
    not, it will be created. If you want to replace it from scratch, delete it
    prior to calling this function. (default is computed as dirPath + ".zip")

    includeDirInZip -- boolean indicating whether the top level directory should
    be included in the archive or omitted. (default True)

    """
    if not zipFilePath:
        zipFilePath = dirPath + ".zip"
    if not os.path.isdir(dirPath):
        raise OSError("dirPath argument must point to a directory. "
            "'%s' does not." % dirPath)
    parentDir, dirToZip = os.path.split(dirPath)
    #Little nested function to prepare the proper archive path
    def trimPath(path):
        archivePath = path.replace(parentDir, "", 1)
        if parentDir:
            archivePath = archivePath.replace(os.path.sep, "", 1)
        if not includeDirInZip:
            archivePath = archivePath.replace(dirToZip + os.path.sep, "", 1)
        return os.path.normcase(archivePath)
        
    outFile = zipfile.ZipFile(zipFilePath, "w",
        compression=zipfile.ZIP_DEFLATED)
    for (archiveDirPath, dirNames, fileNames) in os.walk(dirPath):
        for fileName in fileNames:
            filePath = os.path.join(archiveDirPath, fileName)
            outFile.write(filePath, trimPath(filePath))
        #Make sure we get empty directories as well
        if not fileNames and not dirNames:
            zipInfo = zipfile.ZipInfo(trimPath(archiveDirPath) + "/")
            #some web sites suggest doing
            #zipInfo.external_attr = 16
            #or
            #zipInfo.external_attr = 48
            #Here to allow for inserting an empty directory.  Still TBD/TODO.
            outFile.writestr(zipInfo, "")
    outFile.close()
    print ("ZIP file and its content ready at " + zipFilePath)

if __name__ == "__main__":
    if (len(sys.argv) == 1):
        print("Usage: python ChannelFinder.py <path/to/your/spreadsheets/folder>")
    else:
        sheets_folder = sys.argv[1]
        if not os.path.exists(sheets_folder + '/output'):
            os.makedirs(sheets_folder + '/output')
        for file in os.listdir(sheets_folder):
            # traverse the folder not hidden file, no temporary xlsx file
            if file.endswith(".xlsx") and (not (file.startswith('~'))) and (not (file.startswith('.'))):
                process_workbook(sheets_folder + "/" + file)
        zipdir(sheets_folder + '/output', sheets_folder + '/output.zip', True)
